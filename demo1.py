from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from pandas import DataFrame
import requests
from PIL import Image
import os
import cv2 as cv
from pyzbar.pyzbar import decode
import requests as r
from openpyxl import load_workbook
import ocr
import re



def pyzbarParseQRCode(filePath):
    img = cv.imread(filePath)#读取二维码图片
    texts = decode(img)#解码验证码图片
    for text in texts:#遍历解码数据
        qrInfo = text.data.decode("utf-8")#将内容解码成指定格式
        print(qrInfo)#打印
        return qrInfo

def get_title_content(url,browser):
  # 隐式等待3秒
  browser.implicitly_wait(5)
  browser.get(url)
  title = ''
  content = ''
  mession_detail_list = browser.find_elements(by=By.CLASS_NAME, value='mession-detail-title')
  if len(mession_detail_list) < 1:
      print('内容已被删除')
      title = '内容已被删除'
      return title,content

  title = browser.find_elements(by=By.CLASS_NAME, value='mession-detail-title')[0].text

  print('==============题目===============')
  print(title)

  try:

      section_list = browser.find_elements(by=By.CSS_SELECTOR, value='section')
      video_list = browser.find_elements(by=By.TAG_NAME, value='video')
      img_list = browser.find_elements(by=By.TAG_NAME, value='img')
      content_list = browser.find_elements(by=By.CLASS_NAME, value='mession-detail-content')
      if len(section_list) > 0:
          content = section_list[0].text
      elif len(video_list) > 0:
          content = video_list[0].get_attribute('src')
      elif len(img_list) > 0:
          content = img_list[0].get_attribute("src")
      elif len(content_list) > 0:
          content = content_list[0].text
      print('==============内容===============')
      print(content)

  except Exception as e:
      print(str(e))
      print(url)



  return title,content

def parse_excel(file_path,browser):
      title_list = []
      content_list = []
      chanhou_data = pd.read_excel(file_path)
      print(chanhou_data.head())
      chanhou_href_list = chanhou_data['链接']
      for chanhou_href in chanhou_href_list:
         print(chanhou_href)
         title = ''
         content = ''
         if pd.isna(chanhou_href) == False:
             # 获取网页内容
             title,content = get_title_content(chanhou_href,browser)
         title_list.append(title)
         content_list.append(content)

      chanhou_data['题目'] = title_list
      chanhou_data['内容'] = content_list
      DataFrame(chanhou_data).to_excel(file_path, sheet_name='Sheet1', columns=['链接'],index=False, header=True)


def parse_xuanjiao_excel(file_path,browser):
  title_list = []
  content_list = []
  chanhou_data = pd.read_excel(file_path, sheet_name=None)
  image_num = 0
  for k, v in chanhou_data.items():

      # v = v.to_dict(orient=‘records’)
      if k in ['总揽表']:
          continue
      if v.empty == True:
          continue

      print("============= image_num :"+ str(image_num))

      title_list = []
      content_list = []
      addr_list = ['']
      content_num = len(v['主题'])-1
      while content_num > 0:
          image_num = image_num + 1
          filePath = "./images/image"+str(image_num)+".png"
          addr = pyzbarParseQRCode(filePath)
          addr_list.append(addr)
          content_num = content_num - 1

      if image_num <= 530:
          continue


      for addr_href in addr_list:
          title = ''
          content = ''
          if pd.isna(addr_href) == False and  addr_href != '':
              # 获取网页内容
              title, content = get_title_content(addr_href, browser)
          title_list.append(title)
          content_list.append(content)

      df = pd.DataFrame({'链接': addr_list,'题目':title_list,'内容':content_list})
      book = load_workbook(file_path)  # 加载原有的数据到Workbook
      with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
          writer.book = book
          writer.sheets = {ws.title: ws for ws in book.worksheets}
          max_column = writer.sheets[k].max_column
          df.to_excel(writer, sheet_name=k,index=False, startrow=1,startcol=max_column,header=False)
          writer.save()


def download_url(url,download_path,file_name):
    try:
        print('准备下载:'+url)

        proxy = get_ip_proxy()

        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }

        try:
            response = requests.get('http://httpbin.org/get', proxies=proxies)
            print(response.text)
        except requests.exceptions.ConnectionError as e:
            print('Error', e.args)

        response=requests.get(url)
        data=response.content
        if data:
            # file_path='{}/{}.{}'.format(os.getcwd(),md5(data).hexdigest(),'mp4')
            file_path = download_path + file_name
            print('文件为:'+file_path)
            if not os.path.exists(file_path):
                with open(file_path,'wb')as f:
                    f.write(data)
                    f.close()
                    print('下载成功:'+url)
    except Exception as e:
        print('下载失败')
        print(e)


def download_video_and_ocr(file_path):
  chanhou_data = pd.read_excel(file_path, sheet_name=None)

  for k, v in chanhou_data.items():

      # v = v.to_dict(orient=‘records’)
      if k in ['总揽表']:
          continue
      if v.empty == True:
          continue

      title_list = v['题目']
      content_list = v['内容']
      analysis_content_list = []

      i = 0

      for content in content_list:
          title = title_list[i]
          print(content)
          if pd.isna(content) == True:
              content = ''
          elif re.match(r'^https?:\/\/(.+\/)+.+(\.(gif|png|jpg|jpeg|webp|svg|psd|bmp|tif))$', content):
              print("图片链接:"+content)
              # 阿里云OCR识别：https://ai.aliyun.com/ocr
              download_url(content,'/home/mocuili/github/sp1/ocr_images/',title+'.jpg')
              try:
                content = ocr.Sample.main('/home/mocuili/github/sp1/ocr_images/' + title+'.jpg')
              except Exception:
                content = "图片过长"

          elif  re.match(r'^https?:\/\/(.+\/)+.+(\.(swf|avi|flv|mpg|rm|mov|wav|asf|3gp|mkv|rmvb|mp4))$', content):
              print("视频链接:" + content)
              download_url(content,'/home/mocuili/github/sp1/videos/',title+'.mp4')
              content = ''
          else:
              content = ''

          analysis_content_list.append(content)
          i += 1

      df = pd.DataFrame({'ocr识别': analysis_content_list})
      book = load_workbook(file_path)  # 加载原有的数据到Workbook
      with pd.ExcelWriter(file_path,engine='openpyxl') as writer:
          writer.book = book
          writer.sheets = {ws.title: ws for ws in book.worksheets}
          max_column = writer.sheets[k].max_column
          df.to_excel(writer, sheet_name=k,index=False, startrow=1,startcol=max_column,header=False)
          writer.save()



  # print(content)

def get_proxy():
    try:
        PROXY_POOL_URL = 'http://localhost:5555/random'
        response = requests.get(PROXY_POOL_URL)
        if response.status_code == 200:
            return response.text
    except ConnectionError:
        return None

# 通过代理池获取代理IP
# 代理池github地址：https://github.com/Python3WebSpider/ProxyPool
def get_ip_proxy():

  # proxy = get_proxy()
  # 翻墙VPN的代理IP，如果不用代理池的话，就用这个IP
  proxy = '127.0.0.1:41091'
  return proxy

  # proxy = '127.0.0.1:41091'
  #
  # proxies = {
  #     'http': 'http://' + proxy,
  #     'https': 'https://' + proxy
  # }
  #
  # response = requests.get('https://httpbin.org/get')




try:
  chrome_options = webdriver.ChromeOptions()
  # chrome_options.add_argument('--proxy-server=' + get_ip_proxy())
  browser = webdriver.Chrome(options=chrome_options)
  browser.get('https://httpbin.org/get')

  # 解析excel
  parse_excel(r'/home/mocuili/github/sp1/产后42天SOP.xlsx',browser)
  parse_excel(r'/home/mocuili/github/sp1/孕产增值服务.xlsx',browser)
  parse_xuanjiao_excel(r'/home/mocuili/github/sp1/宣教库统计持续更新0701xlsx(1).xlsx',browser)

  # 图片OCR识别和视频下载
  download_video_and_ocr(r'/home/mocuili/github/sp1/产后42天SOP.xlsx')
  download_video_and_ocr(r'/home/mocuili/github/sp1/孕产增值服务.xlsx')
  download_video_and_ocr(r'/home/mocuili/github/sp1/宣教库统计持续更新0701xlsx(1).xlsx')





except Exception as e:
    print(str(e))
# finally:
  # browser.close()



